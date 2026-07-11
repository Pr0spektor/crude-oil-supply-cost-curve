"""Build model.xlsx — interactive breakeven benchmarking workbook.

Sheets: Outputs (live SUMIF), Data (editable nodes), CostCurve (sorted + chart),
Scenarios (price x cost-inflation grid), MonteCarlo (P10/P50/P90 by price).
Excel and the Python engine agree (verified by headless LibreOffice recalculation).

Run:  python src/build_workbook.py
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from data import (SEGMENTS, DEMAND_MMBD, PRICE_SCENARIOS, COST_INFLATION_SCENARIOS)
from costcurve import (build_curve, total_supply, uneconomic_volume,
                       monte_carlo_at_risk, apply_cost_inflation, marginal_barrel)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

NAVY, LIGHT, WHITE = "1F3A5F", "E8EEF5", "FFFFFF"
head = Font(bold=True, color=WHITE, size=10)
navyf = Font(bold=True, color=NAVY, size=12)
bold = Font(bold=True, size=10)
fill_head = PatternFill("solid", fgColor=NAVY)
fill_lite = PatternFill("solid", fgColor=LIGHT)
yellow = PatternFill("solid", fgColor="FFF6CC")
right = Alignment(horizontal="right")
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _title(ws, t):
    ws["A1"] = t; ws["A1"].font = Font(bold=True, color=NAVY, size=14)


def build_data(ws):
    _title(ws, "Supply nodes (edit yellow cells) — crude + condensate")
    hdr = ["Node", "Region", "Production (mmb/d)", "Breakeven full ($/bbl)",
           "Cash cost ($/bbl)", "BE low", "BE high", "Source"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(3, j, h); c.font = head; c.fill = fill_head
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    r = 4
    for s in SEGMENTS:
        lo, hi = s.band()
        ws.cell(r, 1, s.name); ws.cell(r, 2, s.region)
        for col, val in ((3, s.production_mmbd), (4, s.breakeven_full), (5, s.cash_cost)):
            cc = ws.cell(r, col, val); cc.fill = yellow; cc.border = border; cc.alignment = right
        ws.cell(r, 6, round(lo, 1)).alignment = right
        ws.cell(r, 7, round(hi, 1)).alignment = right
        ws.cell(r, 8, s.source).font = Font(size=8, color="777777")
        r += 1
    last = r - 1
    ws.cell(r + 1, 1, "TOTAL").font = bold
    ws.cell(r + 1, 3, f"=SUM(C4:C{last})").font = bold; ws.cell(r + 1, 3).alignment = right
    for j, w in enumerate([28, 14, 16, 16, 13, 9, 9, 16], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return 4, last


def build_curve_sheet(ws):
    _title(ws, "Supply cost curve (sorted, full-cycle)")
    hdr = ["Rank", "Node", "Production", "Breakeven ($/bbl)", "Cum. end (mmb/d)"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(3, j, h); c.font = head; c.fill = fill_head
    for i, rr in enumerate(build_curve(SEGMENTS, "full"), start=4):
        ws.cell(i, 1, i - 3); ws.cell(i, 2, rr["name"])
        ws.cell(i, 3, rr["production_mmbd"]).alignment = right
        ws.cell(i, 4, rr["cost"]).alignment = right
        ws.cell(i, 5, round(rr["cum_end"], 1)).alignment = right
    last = 3 + len(SEGMENTS)
    ch = BarChart(); ch.type = "col"; ch.title = "Breakeven by node (cheapest first)"
    ch.y_axis.title = "$/bbl"; ch.height = 9; ch.width = 24; ch.legend = None
    ch.add_data(Reference(ws, min_col=4, min_row=3, max_row=last), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=4, max_row=last))
    ws.add_chart(ch, "G3")
    for j, w in enumerate([6, 28, 12, 16, 14], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def build_outputs(ws, a0, a1):
    _title(ws, "Breakeven analysis — live (change the oil-price cell)")
    A = f"Data!$C${a0}:$C${a1}"; F = f"Data!$D${a0}:$D${a1}"; C = f"Data!$E${a0}:$E${a1}"
    ws["A3"] = "Oil price ($/bbl)"; ws["A3"].font = bold
    pc = ws["B3"]; pc.value = PRICE_SCENARIOS["base_65"]; pc.fill = yellow; pc.border = border; pc.alignment = right
    rows = [
        ("Total supply (mmb/d)", f"=SUM({A})", '0.0'),
        ("New-investment at risk — full > price (mmb/d)", f'=SUMIF({F},">"&B3,{A})', '0.0'),
        ("Economic — full <= price (mmb/d)", f'=SUMIF({F},"<="&B3,{A})', '0.0'),
        ("Shut-in risk — cash > price (mmb/d)", f'=SUMIF({C},">"&B3,{A})', '0.0'),
        ("At-risk share (full-cycle)", f'=SUMIF({F},">"&B3,{A})/SUM({A})', '0.0%'),
    ]
    r = 5
    ws.cell(r, 1, "Metric").font = head; ws.cell(r, 1).fill = fill_head
    ws.cell(r, 2, "Value").font = head; ws.cell(r, 2).fill = fill_head
    r += 1
    for label, formula, fmt in rows:
        ws.cell(r, 1, label).font = bold
        cell = ws.cell(r, 2, formula); cell.number_format = fmt
        cell.alignment = right; cell.font = navyf; cell.fill = fill_lite
        r += 1
    ws.cell(r + 1, 1, "Live via SUMIF. Full-cycle = new-investment view; cash cost = shut-in view. "
                      "Economic (not fiscal) breakevens — see METHODS.").font = \
        Font(italic=True, size=8, color="888888")
    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 14


def build_scenarios(ws):
    _title(ws, "Scenario grid — at-risk volume (mmb/d, full-cycle)")
    prices = [PRICE_SCENARIOS[k] for k in ("stress_40", "base_65", "high_85")]
    ws.cell(3, 1, "Cost inflation \\ Oil price").font = head; ws.cell(3, 1).fill = fill_head
    for j, p in enumerate(prices, 2):
        c = ws.cell(3, j, f"${p:.0f}"); c.font = head; c.fill = fill_head; c.alignment = right
    r = 4
    for name, infl in COST_INFLATION_SCENARIOS.items():
        segs = apply_cost_inflation(SEGMENTS, infl)
        ws.cell(r, 1, name).font = bold
        for j, p in enumerate(prices, 2):
            v = uneconomic_volume(segs, p, "full")["at_risk_mmbd"]
            ws.cell(r, j, round(v, 1)).alignment = right
        r += 1
    ws.column_dimensions["A"].width = 22
    for j in range(2, 5):
        ws.column_dimensions[get_column_letter(j)].width = 10


def build_montecarlo(ws):
    _title(ws, "Monte-Carlo — at-risk volume distribution (10k draws, seeded)")
    hdr = ["Oil price ($/bbl)", "Point estimate", "Mean", "P10", "P50", "P90"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(3, j, h); c.font = head; c.fill = fill_head; c.alignment = Alignment(wrap_text=True, horizontal="center")
    r = 4
    for p in (40, 50, 55, 60, 65, 70, 80):
        pt = uneconomic_volume(SEGMENTS, p, "full")["at_risk_mmbd"]
        mc = monte_carlo_at_risk(SEGMENTS, p, n=10000, seed=42)
        vals = [p, pt, mc["mean_at_risk_mmbd"], mc["p10"], mc["p50"], mc["p90"]]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, round(v, 1)); cell.alignment = right
        r += 1
    for j in range(1, 7):
        ws.column_dimensions[get_column_letter(j)].width = 13


def main():
    wb = Workbook()
    ws_out = wb.active; ws_out.title = "Outputs"
    ws_data = wb.create_sheet("Data"); a0, a1 = build_data(ws_data)
    build_outputs(ws_out, a0, a1)
    build_curve_sheet(wb.create_sheet("CostCurve"))
    build_scenarios(wb.create_sheet("Scenarios"))
    build_montecarlo(wb.create_sheet("MonteCarlo"))
    out = os.path.join(ROOT, "model.xlsx"); wb.save(out)
    print("Wrote", out, "| supply %.1f mmb/d" % total_supply(SEGMENTS))


if __name__ == "__main__":
    main()
